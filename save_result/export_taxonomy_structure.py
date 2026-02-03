#!/usr/bin/env python
"""
Export integrated taxonomy structure from all dimensions to Excel
"""
import json
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))
from prompts import dimension_definitions

def parse_taxonomy_tree(node, dimension, parent_path="", level=0, rows=None):
    """Recursively parse taxonomy tree and create table rows"""
    if rows is None:
        rows = []
    
    label = node.get('label', '')
    description = node.get('description', '')
    node_level = node.get('level', level)
    source = node.get('source', 'Initial')
    paper_count = len(node.get('paper_ids', []))
    
    # Build full path
    current_path = f"{parent_path}/{label}" if parent_path else label
    
    # Add current node
    rows.append({
        'Dimension': dimension,
        'Level': node_level,
        'Label': label,
        'Full_Path': current_path,
        'Description': description,
        'Source': source,
        'Paper_Count': paper_count,
        'Indent': '  ' * node_level + label  # For visual hierarchy
    })
    
    # Recursively process children
    children = node.get('children', [])
    if isinstance(children, dict):
        for child_label, child_node in children.items():
            parse_taxonomy_tree(child_node, dimension, current_path, node_level + 1, rows)
    elif isinstance(children, list):
        for child_node in children:
            parse_taxonomy_tree(child_node, dimension, current_path, node_level + 1, rows)
    
    return rows

def load_all_taxonomies(data_dir):
    """Load all taxonomy JSON files"""
    # Load dimensions dynamically from prompts.py
    dimensions = list(dimension_definitions.keys())
    
    all_rows = []
    dim_summaries = []
    
    for dim in dimensions:
        json_file = data_dir / f'final_taxo_{dim}.json'
        if json_file.exists():
            print(f"Loading {json_file.name}...")
            with open(json_file, 'r', encoding='utf-8') as f:
                taxonomy = json.load(f)
            
            # Parse tree
            rows = parse_taxonomy_tree(taxonomy, dim)
            all_rows.extend(rows)
            
            # Summary statistics
            total_nodes = len(rows)
            total_papers = sum(r['Paper_Count'] for r in rows)
            max_level = max(r['Level'] for r in rows) if rows else 0
            
            dim_summaries.append({
                'Dimension': dim,
                'Total_Nodes': total_nodes,
                'Total_Papers': total_papers,
                'Max_Depth': max_level
            })
            
            print(f"  - {total_nodes} nodes, {total_papers} papers, max depth: {max_level}")
        else:
            print(f"Warning: {json_file.name} not found")
    
    return all_rows, dim_summaries

def create_excel_report(data_dir, all_rows, dim_summaries):
    """Create Excel file with multiple sheets"""
    output_file = data_dir / 'integrated_taxonomy_structure.xlsx'
    
    # Create workbook
    wb = Workbook()
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    summary_df = pd.DataFrame(dim_summaries)
    for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_summary.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    # Auto-adjust column widths
    for column in ws_summary.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws_summary.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    # Sheet 2: All Dimensions (Integrated)
    ws_all = wb.create_sheet("All_Dimensions")
    
    df_all = pd.DataFrame(all_rows)
    # Reorder columns
    df_all = df_all[['Dimension', 'Level', 'Indent', 'Label', 'Description', 'Source', 'Paper_Count', 'Full_Path']]
    
    for r_idx, row in enumerate(dataframe_to_rows(df_all, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_all.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            else:
                # Color code by level
                level = df_all.iloc[r_idx - 2]['Level']
                if level == 0:
                    cell.fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
                    cell.font = Font(bold=True)
                elif level == 1:
                    cell.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
    
    # Auto-adjust column widths
    for column in ws_all.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws_all.column_dimensions[column[0].column_letter].width = min(max_length + 2, 60)
    
    # Sheets 3-N: Individual Dimensions (dynamically generated)
    dimensions = [(dim, dim[:30].replace('_', ' ').title()[:31]) for dim in dimension_definitions.keys()]
    
    for dim_full, dim_short in dimensions:
        ws_dim = wb.create_sheet(dim_short)
        
        df_dim = df_all[df_all['Dimension'] == dim_full].copy()
        df_dim = df_dim[['Level', 'Indent', 'Label', 'Description', 'Source', 'Paper_Count', 'Full_Path']]
        
        for r_idx, row in enumerate(dataframe_to_rows(df_dim, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_dim.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:  # Header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                else:
                    # Color code by level
                    level = df_dim.iloc[r_idx - 2]['Level']
                    if level == 0:
                        cell.fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
                        cell.font = Font(bold=True)
                    elif level == 1:
                        cell.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws_dim.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            ws_dim.column_dimensions[column[0].column_letter].width = min(max_length + 2, 60)
    
    # Save workbook
    wb.save(output_file)
    print(f"\n✓ Excel file saved: {output_file}")
    
    return output_file

def main():
    data_dir = Path('/root/taxo/taxoadapt/datasets/posco')
    
    print("=" * 70)
    print("Integrated Taxonomy Structure Export")
    print("=" * 70)
    
    # Load all taxonomies
    print("\n[1/2] Loading taxonomy structures...")
    all_rows, dim_summaries = load_all_taxonomies(data_dir)
    
    # Create Excel report
    print("\n[2/2] Creating Excel report...")
    output_file = create_excel_report(data_dir, all_rows, dim_summaries)
    
    # Print final statistics
    print("\n" + "=" * 70)
    print("Export Complete!")
    print("=" * 70)
    print(f"Total nodes across all dimensions: {len(all_rows)}")
    print(f"Output file: {output_file}")
    print("\nSheets created:")
    print("  1. Summary - Overview of all dimensions")
    print("  2. All_Dimensions - Integrated view")
    print("  3-6. Individual dimension sheets")

if __name__ == "__main__":
    main()
