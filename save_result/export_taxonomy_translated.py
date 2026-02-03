#!/usr/bin/env python
"""
Export integrated taxonomy structure from all dimensions to Excel (Korean Version with Translation)
"""
import json
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from openai import OpenAI
from tqdm import tqdm
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))
from prompts import dimension_definitions

# Dimension name mapping (English -> Korean)
DIMENSION_NAMES = {
    'raw_material_optimization': '원료 최적화',
    'reduction_efficiency_enhancement': '환원 효율 향상',
    'fuel_and_gas_substitution': '연료 및 가스 대체',
    'blast_furnace_structural_and_energy_efficiency_improvement': '고로 구조 및 에너지 효율 개선'
}

def translate_to_korean(text, client, description=False):
    """Translate English text to Korean using GPT API"""
    if not text or text == 'None':
        return ''
    
    try:
        if description:
            system_prompt = "You are a professional translator specializing in steel manufacturing and metallurgy. Translate the following technical description to natural Korean. Keep technical terms accurate."
        else:
            system_prompt = "You are a professional translator specializing in steel manufacturing and metallurgy. Translate the following technical term to natural Korean. Keep it concise."
        
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": text}
            ],
            temperature=0.3,
            max_tokens=500 if description else 100
        )
        
        return response.choices[0].message.content.strip()
    except Exception as e:
        print(f"Translation error: {e}")
        return text  # Return original if translation fails

def parse_taxonomy_tree(node, dimension, client, parent_path="", level=0, rows=None, cache=None):
    """Recursively parse taxonomy tree and create table rows with translation"""
    if rows is None:
        rows = []
    if cache is None:
        cache = {}
    
    label = node.get('label', '')
    description = node.get('description', '')
    node_level = node.get('level', level)
    source = node.get('source', 'Initial')
    paper_count = len(node.get('paper_ids', []))
    
    # Translate label
    if label in cache:
        label_kr = cache[label]
    else:
        label_kr = translate_to_korean(label, client, description=False)
        cache[label] = label_kr
    
    # Translate description
    if description:
        desc_key = f"desc_{description[:50]}"  # Cache key
        if desc_key in cache:
            description_kr = cache[desc_key]
        else:
            description_kr = translate_to_korean(description, client, description=True)
            cache[desc_key] = description_kr
    else:
        description_kr = ''
    
    # Build full path
    current_path = f"{parent_path}/{label}" if parent_path else label
    current_path_kr = f"{parent_path.replace(label.split('/')[-1] if '/' in parent_path else parent_path, cache.get(parent_path.split('/')[-1] if '/' in parent_path else parent_path, parent_path))}/{label_kr}" if parent_path else label_kr
    
    # Add current node
    rows.append({
        '차원': DIMENSION_NAMES.get(dimension, dimension),
        '레벨': node_level,
        '기술명 (영문)': label,
        '기술명 (한글)': label_kr,
        '전체 경로': current_path,
        '설명 (영문)': description if description else '',
        '설명 (한글)': description_kr,
        '출처': source,
        '논문 수': paper_count,
        '계층 구조': '  ' * node_level + label_kr
    })
    
    # Recursively process children
    children = node.get('children', [])
    if isinstance(children, dict):
        for child_label, child_node in children.items():
            parse_taxonomy_tree(child_node, dimension, client, current_path, node_level + 1, rows, cache)
    elif isinstance(children, list):
        for child_node in children:
            parse_taxonomy_tree(child_node, dimension, client, current_path, node_level + 1, rows, cache)
    
    return rows

def load_all_taxonomies(data_dir, client):
    """Load all taxonomy JSON files and translate"""
    # Load dimensions dynamically from prompts.py
    dimensions = list(dimension_definitions.keys())
    
    all_rows = []
    dim_summaries = []
    cache = {}  # Translation cache
    
    for dim in dimensions:
        json_file = data_dir / f'final_taxo_{dim}.json'
        if json_file.exists():
            print(f"\nLoading and translating {json_file.name}...")
            with open(json_file, 'r', encoding='utf-8') as f:
                taxonomy = json.load(f)
            
            # Parse tree with translation
            rows = parse_taxonomy_tree(taxonomy, dim, client, cache=cache)
            all_rows.extend(rows)
            
            # Summary statistics
            total_nodes = len(rows)
            total_papers = sum(r['논문 수'] for r in rows)
            max_level = max(r['레벨'] for r in rows) if rows else 0
            
            dim_summaries.append({
                '차원': DIMENSION_NAMES.get(dim, dim),
                '총 노드 수': total_nodes,
                '총 논문 수': total_papers,
                '최대 깊이': max_level
            })
            
            print(f"  ✓ {total_nodes}개 노드, {total_papers}개 논문, 최대 깊이: {max_level}")
        else:
            print(f"Warning: {json_file.name} not found")
    
    return all_rows, dim_summaries

def create_excel_report(data_dir, all_rows, dim_summaries):
    """Create Excel file with multiple sheets"""
    output_file = data_dir / 'integrated_taxonomy_structure_korean_translated.xlsx'
    
    # Create workbook
    wb = Workbook()
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "요약"
    
    summary_df = pd.DataFrame(dim_summaries)
    for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_summary.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header
                cell.font = Font(bold=True, name='맑은 고딕')
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            else:
                cell.font = Font(name='맑은 고딕')
    
    # Auto-adjust column widths
    for column in ws_summary.columns:
        max_length = max(len(str(cell.value or "")) * 1.5 for cell in column)
        ws_summary.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    # Sheet 2: All Dimensions (Integrated)
    ws_all = wb.create_sheet("전체 차원")
    
    df_all = pd.DataFrame(all_rows)
    # Reorder columns - show Korean first
    df_all = df_all[['차원', '레벨', '계층 구조', '기술명 (한글)', '기술명 (영문)', '설명 (한글)', '설명 (영문)', '출처', '논문 수', '전체 경로']]
    
    for r_idx, row in enumerate(dataframe_to_rows(df_all, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_all.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header
                cell.font = Font(bold=True, name='맑은 고딕')
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            else:
                cell.font = Font(name='맑은 고딕')
                # Color code by level
                level = df_all.iloc[r_idx - 2]['레벨']
                if level == 0:
                    cell.fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
                    cell.font = Font(bold=True, name='맑은 고딕')
                elif level == 1:
                    cell.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
    
    # Auto-adjust column widths
    for column in ws_all.columns:
        max_length = max(len(str(cell.value or "")) * 1.3 for cell in column)
        ws_all.column_dimensions[column[0].column_letter].width = min(max_length + 2, 100)
    
    # Sheets 3-N: Individual Dimensions (dynamically generated)
    # For Translated version, use Korean name from DIMENSION_NAMES if available
    dimensions_list = [(DIMENSION_NAMES.get(dim, dim), dim) for dim in dimension_definitions.keys()]
    
    for dim_korean, dim_english in dimensions_list:
        ws_dim = wb.create_sheet(dim_korean[:31])  # Excel sheet name limit
        
        df_dim = df_all[df_all['차원'] == dim_korean].copy()
        df_dim = df_dim[['레벨', '계층 구조', '기술명 (한글)', '기술명 (영문)', '설명 (한글)', '설명 (영문)', '출처', '논문 수', '전체 경로']]
        
        for r_idx, row in enumerate(dataframe_to_rows(df_dim, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_dim.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:  # Header
                    cell.font = Font(bold=True, name='맑은 고딕')
                    cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                else:
                    cell.font = Font(name='맑은 고딕')
                    # Color code by level
                    if len(df_dim) > 0:
                        level = df_dim.iloc[r_idx - 2]['레벨']
                        if level == 0:
                            cell.fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
                            cell.font = Font(bold=True, name='맑은 고딕')
                        elif level == 1:
                            cell.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws_dim.columns:
            max_length = max(len(str(cell.value or "")) * 1.3 for cell in column)
            ws_dim.column_dimensions[column[0].column_letter].width = min(max_length + 2, 100)
    
    # Save workbook
    wb.save(output_file)
    print(f"\n✓ 한글 번역 엑셀 파일 저장 완료: {output_file}")
    
    return output_file

def main():
    data_dir = Path('/root/taxo/taxoadapt/datasets/posco')
    
    # Initialize OpenAI client
    api_key = os.getenv("OPENAI_API_KEY")

    if not api_key:
        raise ValueError("OPENAI_API_KEY environment variable not set")
    
    client = OpenAI(api_key=api_key)
    
    print("=" * 70)
    print("통합 기술 분류 체계 엑셀 생성 (한글 번역 버전)")
    print("=" * 70)
    
    # Load all taxonomies
    print("\n[1/2] 기술 분류 체계 로딩 및 번역 중...")
    print("(GPT API를 사용하여 기술명과 설명을 한글로 번역합니다)")
    all_rows, dim_summaries = load_all_taxonomies(data_dir, client)
    
    # Create Excel report
    print("\n[2/2] 엑셀 리포트 생성 중...")
    output_file = create_excel_report(data_dir, all_rows, dim_summaries)
    
    # Print final statistics
    print("\n" + "=" * 70)
    print("생성 완료!")
    print("=" * 70)
    print(f"전체 차원의 총 노드 수: {len(all_rows)}")
    print(f"출력 파일: {output_file}")
    print("\n생성된 시트:")
    print("  1. 요약 - 전체 차원 개요")
    print("  2. 전체 차원 - 통합 뷰 (한글/영문 병기)")
    print("  3-6. 개별 차원별 시트 (한글/영문 병기)")
    print("\n※ 기술명과 설명이 GPT를 통해 한글로 번역되었습니다.")

if __name__ == "__main__":
    main()
